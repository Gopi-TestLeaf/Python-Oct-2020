# Step 1: install openpyxl:- pip install openpyxl

# Read data from Excel
import openpyxl

# load the excel_workbook
wb = openpyxl.load_workbook(r"C:\Users\gopir\PycharmProjects\Python_Oct2020\data\ReadData.xlsx")

# return all the sheet names of wb
print(wb.sheetnames)

# return active sheet name of wb
print(wb.active)

# create object for specific sheet
sh = wb['Mentors']

# for row count
print("total rows: ", sh.max_row)

# for column count
print("total count: ", sh.max_column)

# read the data
# Way 1
# for row in range(1, sh.max_row+1):
#     for column in range(1, sh.max_column+1):
#         print(sh.cell(row, column).value)

# Way 2
for row in sh['A1' : 'C4']:
    for cell in row:
        print(cell.value)